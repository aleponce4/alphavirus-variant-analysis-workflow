#!/usr/bin/env python3
"""
export_excel_variants.py
========================
Generates publication-ready, formatted Excel (.xlsx) workbooks from variant calling outputs.

Features:
- Clear, human-understandable column headers
- Short AA format (e.g., T126A, C80=) and Long HGVS format (e.g., p.Thr126Ala, p.Cys80=)
- Single-letter and 3-letter amino acid breakdowns
- Percentage formatting for Variant Allele Frequency (VAF) (e.g., 98.28%)
- Comma-formatted Read Depth (DP) and Quality Scores (QUAL)
- Color-coded highlighting for nonsynonymous mutations (soft amber) and high frequency variants
- Multiple structured sheets:
    1. Executive Summary & Quality Statistics
    2. All Filtered Variants
    3. Nonsynonymous Mutations
    4. Synonymous Mutations
    5. Major Variants (VAF >= 5%)
    6. Minor Quasispecies (VAF < 5%)
- Auto-adjusted column widths, frozen headers, gridlines enabled, and auto-filters.
"""

import sys
import re
import argparse
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

AA_3LETTER = {
    'A': 'Ala', 'C': 'Cys', 'D': 'Asp', 'E': 'Glu', 'F': 'Phe',
    'G': 'Gly', 'H': 'His', 'I': 'Ile', 'K': 'Lys', 'L': 'Leu',
    'M': 'Met', 'N': 'Asn', 'P': 'Pro', 'Q': 'Gln', 'R': 'Arg',
    'S': 'Ser', 'T': 'Thr', 'V': 'Val', 'W': 'Trp', 'Y': 'Tyr',
    '*': 'Ter', 'X': 'Xaa'
}

def parse_aa_change(aa_str: str, consequence: str) -> dict:
    """Parses raw AA change strings into short, long (HGVS), 1-letter, and 3-letter fields."""
    if pd.isna(aa_str) or not str(aa_str).strip() or aa_str == '.':
        return {
            'aa_pos': None,
            'ref_aa_1': '', 'alt_aa_1': '',
            'ref_aa_3': '', 'alt_aa_3': '',
            'aa_short': '', 'aa_long': ''
        }
    
    aa_str = str(aa_str).strip()
    
    # Pattern 1: Missense / Nonsynonymous like '126T>126A', '3K>3E', or '123A>123*'
    m1 = re.match(r'^(\d+)([A-Z\*])>(\d+)([A-Z\*])$', aa_str)
    if m1:
        pos1, ref, pos2, alt = m1.groups()
        ref_3 = AA_3LETTER.get(ref, ref)
        alt_3 = AA_3LETTER.get(alt, alt)
        short_fmt = f"{ref}{pos1}{alt}"
        long_fmt = f"p.{ref_3}{pos1}{alt_3}"
        return {
            'aa_pos': int(pos1),
            'ref_aa_1': ref, 'alt_aa_1': alt,
            'ref_aa_3': ref_3, 'alt_aa_3': alt_3,
            'aa_short': short_fmt, 'aa_long': long_fmt
        }
    
    # Pattern 2: Synonymous like '80C' or '102K'
    m2 = re.match(r'^(\d+)([A-Z\*])$', aa_str)
    if m2:
        pos, ref = m2.groups()
        ref_3 = AA_3LETTER.get(ref, ref)
        short_fmt = f"{ref}{pos}="
        long_fmt = f"p.{ref_3}{pos}="
        return {
            'aa_pos': int(pos),
            'ref_aa_1': ref, 'alt_aa_1': ref,
            'ref_aa_3': ref_3, 'alt_aa_3': ref_3,
            'aa_short': short_fmt, 'aa_long': long_fmt
        }
    
    # Fallback for complex indels or unusual annotations
    return {
        'aa_pos': None,
        'ref_aa_1': '', 'alt_aa_1': '',
        'ref_aa_3': '', 'alt_aa_3': '',
        'aa_short': aa_str, 'aa_long': f"p.{aa_str}" if not aa_str.startswith('p.') else aa_str
    }

def format_dna_change(dna_str: str) -> str:
    """Ensures standard HGVS notation for DNA changes (e.g. c.422A>G)."""
    if pd.isna(dna_str) or not str(dna_str).strip() or dna_str == '.':
        return ""
    s = str(dna_str).strip()
    if not s.startswith('c.'):
        return f"c.{s}"
    return s

def prepare_enhanced_dataframe(tsv_path: Path) -> pd.DataFrame:
    """Reads input variant TSV and builds an expanded, formatted DataFrame."""
    df = pd.read_csv(tsv_path, sep='\t')
    
    parsed_aa = [parse_aa_change(row.get('amino_acid_change'), row.get('consequence')) for _, row in df.iterrows()]
    parsed_aa_df = pd.DataFrame(parsed_aa)
    
    formatted_df = pd.DataFrame()
    formatted_df['Sample ID'] = df['sample_name']
    formatted_df['Days Post Infection (DPI)'] = pd.to_numeric(df['dpi'], errors='coerce')
    formatted_df['Replicate'] = pd.to_numeric(df['replicate'], errors='coerce')
    formatted_df['Treatment Group'] = df['treatment']
    formatted_df['Virus Reference'] = df['CHROM']
    formatted_df['Genome Position (bp)'] = pd.to_numeric(df['POS'], errors='coerce')
    formatted_df['Ref Base'] = df['REF']
    formatted_df['Alt Base'] = df['ALT']
    formatted_df['DNA Change (c.HGVS)'] = df['dna_change'].apply(format_dna_change)
    formatted_df['Gene / Protein'] = df['gene']
    
    # Standardize 'missense' consequence to 'nonsynonymous' to match plots
    formatted_df['Mutation Type'] = df['consequence'].replace({'missense': 'nonsynonymous'})
    
    formatted_df['AA Position'] = parsed_aa_df['aa_pos']
    formatted_df['AA Change (Short)'] = parsed_aa_df['aa_short']
    formatted_df['AA Change (Long / HGVS)'] = parsed_aa_df['aa_long']
    formatted_df['Ref AA (1-letter)'] = parsed_aa_df['ref_aa_1']
    formatted_df['Alt AA (1-letter)'] = parsed_aa_df['alt_aa_1']
    formatted_df['Ref AA (3-letter)'] = parsed_aa_df['ref_aa_3']
    formatted_df['Alt AA (3-letter)'] = parsed_aa_df['alt_aa_3']
    
    # Variant frequency and depth
    formatted_df['Variant Allele Frequency (%)'] = pd.to_numeric(df['AF'], errors='coerce')
    formatted_df['Read Depth (DP)'] = pd.to_numeric(df['DP'], errors='coerce')
    formatted_df['Quality Score (QUAL)'] = pd.to_numeric(df['QUAL'], errors='coerce')
    formatted_df['Filter Status'] = df['FILTER']
    formatted_df['Variant Caller'] = df['caller']
    
    return formatted_df

def style_worksheet(ws: openpyxl.worksheet.worksheet.Worksheet, is_summary: bool = False):
    """Applies clean, professional styles, number formatting, colors, and column width auto-fitting."""
    ws.views.sheetView[0].showGridLines = True
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid") # Deep navy
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    
    nonsyn_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Soft amber
    nonsyn_font = Font(name="Segoe UI", size=10, color="92400E", bold=True)
    
    alt_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    # Header styling
    max_row = ws.max_row
    max_col = ws.max_column
    
    if max_row == 0 or max_col == 0:
        return

    # Enable autofilter & freeze row 1 if not summary
    if not is_summary:
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    
    headers = [cell.value for cell in ws[1]]
    
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 28

    # Style data rows
    for row_idx in range(2, max_row + 1):
        ws.row_dimensions[row_idx].height = 20
        is_even = (row_idx % 2 == 0)
        
        # Check mutation type for highlighting
        mut_type = None
        if 'Mutation Type' in headers:
            mut_col_idx = headers.index('Mutation Type') + 1
            mut_type = ws.cell(row=row_idx, column=mut_col_idx).value
            
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            header_name = headers[col_idx - 1]
            cell.border = thin_border
            cell.font = Font(name="Segoe UI", size=10)
            
            # Default background row striping
            if is_even and not is_summary:
                cell.fill = alt_row_fill
                
            # Formatting by column type
            if header_name in ['Genome Position (bp)', 'AA Position', 'Days Post Infection (DPI)', 'Replicate', 'Read Depth (DP)']:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0'
            elif header_name in ['Variant Allele Frequency (%)', 'Average VAF (%)', '% Total Variants']:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '0.00%'
            elif header_name in ['Quality Score (QUAL)', 'Average Depth (DP)', 'Mean Coverage']:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0.0'
            elif header_name in ['Ref Base', 'Alt Base', 'Ref AA (1-letter)', 'Alt AA (1-letter)', 'Filter Status', 'Variant Caller', 'Mutation Type']:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Highlight Nonsynonymous rows specially in data tables
            if mut_type in ['nonsynonymous', 'missense']:
                if header_name in ['Mutation Type', 'AA Change (Short)', 'AA Change (Long / HGVS)', 'AA Position']:
                    cell.fill = nonsyn_fill
                    cell.font = nonsyn_font
                    
    # Auto-adjust column widths
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        header_val = str(col[0].value or '')
        
        for cell in col:
            val = cell.value
            if val is not None:
                if isinstance(val, float):
                    if 'Frequency' in header_val or '%' in header_val:
                        val_str = f"{val * 100:.2f}%"
                    else:
                        val_str = f"{val:,.2f}"
                else:
                    val_str = str(val)
                max_len = max(max_len, len(val_str))
                
        # Padding
        adjusted_width = max(max_len + 3, len(header_val) + 4, 12)
        ws.column_dimensions[col_letter].width = min(adjusted_width, 40)

def create_summary_sheet(wb: openpyxl.Workbook, df: pd.DataFrame, dataset_name: str):
    """Creates a high-level study overview and quality statistics tab."""
    ws = wb.create_sheet(title="Study Overview & QC", index=0)
    
    title_font = Font(name="Segoe UI", size=16, bold=True, color="1B365D")
    section_font = Font(name="Segoe UI", size=12, bold=True, color="1B365D")
    bold_label_font = Font(name="Segoe UI", size=10, bold=True)
    regular_font = Font(name="Segoe UI", size=10)
    
    ws['A1'] = f"Variant Calling Analysis Overview ({dataset_name})"
    ws['A1'].font = title_font
    
    total_vars = len(df)
    unique_samples = df['Sample ID'].nunique()
    nonsyn_cnt = (df['Mutation Type'].isin(['nonsynonymous', 'missense'])).sum()
    synonymous_cnt = (df['Mutation Type'] == 'synonymous').sum()
    other_cnt = total_vars - nonsyn_cnt - synonymous_cnt
    mean_vaf = df['Variant Allele Frequency (%)'].mean()
    mean_depth = df['Read Depth (DP)'].mean()
    
    # Overview Metrics Table
    ws['A3'] = "Dataset Metrics Overview"
    ws['A3'].font = section_font
    
    metrics = [
        ("Total Variants Called", total_vars, '#,##0'),
        ("Total Samples Analyzed", unique_samples, '#,##0'),
        ("Nonsynonymous Mutations", nonsyn_cnt, '#,##0'),
        ("Synonymous Mutations", synonymous_cnt, '#,##0'),
        ("Other / Indel / Stop Mutations", other_cnt, '#,##0'),
        ("Average Variant Allele Frequency (VAF)", mean_vaf, '0.00%'),
        ("Average Read Depth (DP)", mean_depth, '#,##0.0')
    ]
    
    ws['A4'] = "Metric"
    ws['B4'] = "Value"
    ws['A4'].font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    ws['B4'].font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    ws['A4'].fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    ws['B4'].fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    
    curr_row = 5
    for label, val, num_fmt in metrics:
        ws.cell(row=curr_row, column=1, value=label).font = bold_label_font
        cell_val = ws.cell(row=curr_row, column=2, value=val)
        cell_val.font = regular_font
        cell_val.number_format = num_fmt
        cell_val.alignment = Alignment(horizontal="right")
        curr_row += 1
        
    # Breakdown by Gene
    ws.cell(row=curr_row + 2, column=1, value="Gene Breakdown Summary").font = section_font
    
    gene_summary = df.groupby('Gene / Protein').agg(
        Total_Variants=('Genome Position (bp)', 'count'),
        Nonsynonymous=('Mutation Type', lambda x: (x.isin(['nonsynonymous', 'missense'])).sum()),
        Synonymous=('Mutation Type', lambda x: (x == 'synonymous').sum()),
        Mean_VAF=('Variant Allele Frequency (%)', 'mean'),
        Mean_DP=('Read Depth (DP)', 'mean')
    ).reset_index()
    
    g_start_row = curr_row + 4
    g_headers = ["Gene / Protein", "Total Variants", "Nonsynonymous", "Synonymous", "Average VAF (%)", "Average Depth (DP)"]
    
    for c_idx, h_text in enumerate(g_headers, 1):
        cell = ws.cell(row=g_start_row - 1, column=c_idx, value=h_text)
        cell.fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        cell.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for r_idx, g_row in gene_summary.iterrows():
        row_num = g_start_row + r_idx
        ws.cell(row=row_num, column=1, value=g_row['Gene / Protein']).font = bold_label_font
        
        c2 = ws.cell(row=row_num, column=2, value=g_row['Total_Variants'])
        c2.number_format = '#,##0'
        c2.alignment = Alignment(horizontal="right")
        
        c3 = ws.cell(row=row_num, column=3, value=g_row['Nonsynonymous'])
        c3.number_format = '#,##0'
        c3.alignment = Alignment(horizontal="right")
        
        c4 = ws.cell(row=row_num, column=4, value=g_row['Synonymous'])
        c4.number_format = '#,##0'
        c4.alignment = Alignment(horizontal="right")
        
        c5 = ws.cell(row=row_num, column=5, value=g_row['Mean_VAF'])
        c5.number_format = '0.00%'
        c5.alignment = Alignment(horizontal="right")
        
        c6 = ws.cell(row=row_num, column=6, value=g_row['Mean_DP'])
        c6.number_format = '#,##0.0'
        c6.alignment = Alignment(horizontal="right")
        
    style_worksheet(ws, is_summary=True)

def build_e2_k3e_audit_sheet(wb: openpyxl.Workbook, base_dir: Path, dataset_name: str):
    """Builds a dedicated audit worksheet for E2-K3E (Site 8570) read coverage and allele percentages."""
    manifest_path = base_dir.parent / "config" / "samples_manifest.tsv"
    if not manifest_path.exists():
        manifest_path = Path("variant_analysis/config/samples_manifest.tsv")
        
    if not manifest_path.exists():
        return
        
    manifest = pd.read_csv(manifest_path, sep='\t')
    ds_manifest = manifest[manifest['dataset'] == dataset_name]
    if ds_manifest.empty:
        return

    rows = []
    for idx, row in ds_manifest.iterrows():
        s = str(row['sample_id'])
        s_full = f's{s}' if not s.startswith('s') else s
        dpi = row['dpi']
        treatment = row['treatment']
        
        cov_path = base_dir / dataset_name / 'Coverage' / f'{s_full}_coverage.txt'
        depth_8570 = 0
        if cov_path.exists():
            cov_df = pd.read_csv(cov_path, sep='\t', header=None, names=['chrom', 'pos', 'depth'])
            match = cov_df[cov_df['pos'] == 8570]
            if not match.empty:
                depth_8570 = int(match['depth'].values[0])
                
        site_path = base_dir / dataset_name / 'SNPGenie' / 'output' / 'minfreq_0p01' / s_full / 'site_results.txt'
        a_cnt, c_cnt, g_cnt, t_cnt = 0, 0, 0, 0
        if site_path.exists():
            site_df = pd.read_csv(site_path, sep='\t')
            m = site_df[site_df['site'] == 8570]
            if not m.empty:
                a_cnt = int(m['A'].values[0])
                c_cnt = int(m['C'].values[0])
                g_cnt = int(m['G'].values[0])
                t_cnt = int(m['T'].values[0])
                depth_8570 = int(m['coverage'].values[0])
            else:
                a_cnt = depth_8570
        else:
            a_cnt = depth_8570
            
        ref_pct = (a_cnt / depth_8570) if depth_8570 > 0 else 0.0
        var_pct = (g_cnt / depth_8570) if depth_8570 > 0 else 0.0
        other_cnt = c_cnt + t_cnt
        other_pct = (other_cnt / depth_8570) if depth_8570 > 0 else 0.0
        
        qc_status = 'PASS (>= 1,000x)' if depth_8570 >= 1000 else 'Low Coverage (< 1,000x)'
        
        rows.append({
            'Sample ID': s_full,
            'DPI': dpi,
            'Treatment': treatment,
            'Site 8570 Read Depth': depth_8570,
            'Ref Allele (A - Lysine) Count': a_cnt,
            'Ref Allele (A) %': ref_pct,
            'Variant Allele (G - E2-K3E) Count': g_cnt,
            'Variant Allele (G - E2-K3E) %': var_pct,
            'Other Allele Count': other_cnt,
            'Other Allele %': other_pct,
            'QC / Filter Status': qc_status
        })

    audit_df = pd.DataFrame(rows)
    if audit_df.empty:
        return

    sheet_title = "VEEV E2-K3E Audit" if "master" in str(wb).lower() or len(wb.sheetnames) > 4 else "E2-K3E Audit (Site 8570)"
    ws = wb.create_sheet(title=sheet_title)
    for r in dataframe_to_rows(audit_df, index=False, header=True):
        ws.append(r)
        
    style_worksheet(ws)
    
    # Custom cell number formatting for percentages and integer depth
    for r_idx in range(2, ws.max_row + 1):
        for col_idx in [4, 5, 7, 9]:
            ws.cell(row=r_idx, column=col_idx).number_format = '#,##0'
        for col_idx in [6, 8, 10]:
            ws.cell(row=r_idx, column=col_idx).number_format = '0.00%'

def save_workbook_safely(wb: openpyxl.Workbook, output_path: Path):
    """Saves workbook and handles PermissionError gracefully if open in Excel."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(output_path)
        print(f"  [OK] Saved formatted Excel workbook to: {output_path}")
    except PermissionError:
        alt_path = output_path.parent / f"{output_path.stem}_updated{output_path.suffix}"
        wb.save(alt_path)
        print(f"  [WARNING] File '{output_path.name}' is currently open in Excel.")
        print(f"  [OK] Saved updated copy to: {alt_path}")

def export_dataset_to_excel(tsv_path: Path, output_excel_path: Path, dataset_name: str):
    """Builds a complete, formatted Excel workbook with multiple tabs for a given dataset."""
    print(f"Processing dataset '{dataset_name}' from: {tsv_path}")
    df = prepare_enhanced_dataframe(tsv_path)
    
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # 1. Study Overview & QC Stats Tab
    create_summary_sheet(wb, df, dataset_name)
    
    # 2. All Variants Tab
    ws_all = wb.create_sheet(title="All Variants")
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_all.append(r)
    style_worksheet(ws_all)
    
    # 3. Nonsynonymous Tab
    df_nonsyn = df[df['Mutation Type'].isin(['nonsynonymous', 'missense'])]
    if not df_nonsyn.empty:
        ws_nonsyn = wb.create_sheet(title="Nonsynonymous Mutations")
        for r in dataframe_to_rows(df_nonsyn, index=False, header=True):
            ws_nonsyn.append(r)
        style_worksheet(ws_nonsyn)
        
    # 4. Synonymous Tab
    df_syn = df[df['Mutation Type'] == 'synonymous']
    if not df_syn.empty:
        ws_syn = wb.create_sheet(title="Synonymous Mutations")
        for r in dataframe_to_rows(df_syn, index=False, header=True):
            ws_syn.append(r)
        style_worksheet(ws_syn)
        
    # 5. Major Variants (VAF >= 5%)
    df_major = df[df['Variant Allele Frequency (%)'] >= 0.05]
    if not df_major.empty:
        ws_major = wb.create_sheet(title="Major Variants (VAF >= 5%)")
        for r in dataframe_to_rows(df_major, index=False, header=True):
            ws_major.append(r)
        style_worksheet(ws_major)

    # 6. Minor Variants (VAF < 5%)
    df_minor = df[df['Variant Allele Frequency (%)'] < 0.05]
    if not df_minor.empty:
        ws_minor = wb.create_sheet(title="Minor Variants (VAF < 5%)")
        for r in dataframe_to_rows(df_minor, index=False, header=True):
            ws_minor.append(r)
        style_worksheet(ws_minor)

    # 7. Dedicated E2-K3E Audit Tab for VEEV
    if "veev" in dataset_name.lower():
        build_e2_k3e_audit_sheet(wb, tsv_path.parent.parent.parent, dataset_name)

    save_workbook_safely(wb, output_excel_path)

def main():
    parser = argparse.ArgumentParser(description="Export variant TSVs to beautifully formatted Excel workbooks.")
    parser.add_argument("--results-dir", help="Path to base results directory", default="variant_analysis/results")
    args = parser.parse_args()
    
    base_dir = Path(args.results_dir)
    
    datasets = ["mouse_eeev", "mouse_veev", "rat_veev"]
    
    all_dfs = []
    
    for ds in datasets:
        tsv_path = base_dir / ds / "tables" / "combined_filtered_variants.tsv"
        excel_out = base_dir / ds / "tables" / f"{ds}_formatted_variants.xlsx"
        
        if tsv_path.exists():
            export_dataset_to_excel(tsv_path, excel_out, ds)
            df = prepare_enhanced_dataframe(tsv_path)
            df['Dataset'] = ds
            all_dfs.append(df)
        else:
            print(f"Skipping dataset '{ds}': TSV not found at {tsv_path}")
            
    # If multiple datasets present, create a Master Combined Excel File
    if len(all_dfs) > 1:
        combined_df = pd.concat(all_dfs, ignore_index=True)
        cols = ['Dataset'] + [c for c in combined_df.columns if c != 'Dataset']
        combined_df = combined_df[cols]
        
        master_excel_out = base_dir / "all_datasets_formatted_variants_master.xlsx"
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        create_summary_sheet(wb, combined_df, "All Alphavirus Datasets Combined")
        
        ws_master = wb.create_sheet(title="All Datasets Variants")
        for r in dataframe_to_rows(combined_df, index=False, header=True):
            ws_master.append(r)
        style_worksheet(ws_master)
        
        # Master Nonsynonymous Tab
        df_nonsyn_all = combined_df[combined_df['Mutation Type'].isin(['nonsynonymous', 'missense'])]
        if not df_nonsyn_all.empty:
            ws_n_master = wb.create_sheet(title="Master Nonsynonymous")
            for r in dataframe_to_rows(df_nonsyn_all, index=False, header=True):
                ws_n_master.append(r)
            style_worksheet(ws_n_master)
            
        # Master Synonymous Tab
        df_syn_all = combined_df[combined_df['Mutation Type'] == 'synonymous']
        if not df_syn_all.empty:
            ws_s_master = wb.create_sheet(title="Master Synonymous")
            for r in dataframe_to_rows(df_syn_all, index=False, header=True):
                ws_s_master.append(r)
            style_worksheet(ws_s_master)
            
        # Master Major Variants (VAF >= 5%)
        df_major_all = combined_df[combined_df['Variant Allele Frequency (%)'] >= 0.05]
        if not df_major_all.empty:
            ws_maj_master = wb.create_sheet(title="Master Major (VAF >= 5%)")
            for r in dataframe_to_rows(df_major_all, index=False, header=True):
                ws_maj_master.append(r)
            style_worksheet(ws_maj_master)

        # Master Minor Variants (VAF < 5%)
        df_minor_all = combined_df[combined_df['Variant Allele Frequency (%)'] < 0.05]
        if not df_minor_all.empty:
            ws_min_master = wb.create_sheet(title="Master Minor (VAF < 5%)")
            for r in dataframe_to_rows(df_minor_all, index=False, header=True):
                ws_min_master.append(r)
            style_worksheet(ws_min_master)

        # VEEV E2-K3E Audit Tab in Master
        build_e2_k3e_audit_sheet(wb, base_dir, "mouse_veev")
            
        save_workbook_safely(wb, master_excel_out)

if __name__ == "__main__":
    main()
